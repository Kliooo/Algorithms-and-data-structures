import os
import string
import subprocess
import hashlib
import random
import pandas as pd
import tkinter as tk
from openpyxl import load_workbook
from tkinter import filedialog, messagebox
from openpyxl.styles import Font, Alignment

selectedEncryptionFile = None

def selectEncryptedFile():
    selectedFilePath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if selectedFilePath:
        fileName = os.path.basename(selectedFilePath)
        fileLabel.set(f"Выбранный для деобезличивания файл: {fileName}")
        btnDeobfuscate.config(state=tk.NORMAL)
    else:
        fileLabel.set("Выбранный для деобезличивания файл: Файл не выбран")

def selectFileForEncryption():
    global selectedEncryptionFile
    selectedEncryptionFile = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    
    if selectedEncryptionFile:
        fileName = os.path.basename(selectedEncryptionFile)
        encryptFileLabel.set(f"Выбранный для зашифровки файл: {fileName}")
        sha1Button.config(state=tk.NORMAL)
        blake2bButton.config(state=tk.NORMAL)
        MD5Button.config(state=tk.NORMAL)
    else:
        encryptFileLabel.set("Выбранный для зашифровки файл: Файл не выбран")

def runHashcat(filePath):
    currentDir = os.path.dirname(os.path.abspath(__file__))
    outputFilePath = os.path.join(currentDir, "output.txt")
    hashcatExe = os.path.join(currentDir, "hashcat-6.2.6", "hashcat.exe")

    result = extractPasswordHashes(filePath)
    hashesFilePath = result[1]

    hashcatCommand = [
        hashcatExe, "-a", "3", "-m", "0", "-o", outputFilePath, hashesFilePath, "?d?d?d?d?d?d?d?d?d?d?d"
    ]
    
    result = subprocess.run(hashcatCommand, cwd=os.path.join(currentDir, "hashcat-6.2.6"), capture_output=True, text=True)
    print(result.stdout)

    return outputFilePath

def extractPasswordHashes(filePath):
    df = pd.read_excel(filePath)
    passwordHashes = df.iloc[:, 0].dropna().values
    currentDir = os.path.dirname(os.path.abspath(__file__))
    hashesFilePath = os.path.join(currentDir, "hashcat-6.2.6", "hashes_for_hashcat.txt")

    with open(hashesFilePath, "w") as f:
        for h in passwordHashes:
            f.write(f"{h}\n")
    return len(passwordHashes), hashesFilePath

def applyDecryptionSalt(filePath, decryptedFilePath):
    phonesFilePath = os.path.join(os.path.dirname(os.path.abspath(__file__)), "phones.xlsx")

    df = pd.read_excel(filePath)
    knownNumbers = df.iloc[:, 2].dropna().astype(int).tolist()

    decryptedDict = {}
    with open(decryptedFilePath, 'r') as f:
        for line in f:
            hashVal, decryptedNumber = line.strip().split(':')
            decryptedDict[hashVal] = decryptedNumber

    def computeSalts(knownNumbers, decryptedDict):
        salts = set()
        for decrypted in decryptedDict.values():
            try:
                salt = int(decrypted) - knownNumbers[0]
                if salt < 0:
                    continue
                match = True
                for knownNum in knownNumbers:
                    if str(knownNum + salt) not in decryptedDict.values():
                        match = False
                        break
                if match:
                    salts.add(salt)
            except ValueError:
                continue
        return list(salts)

    salts = computeSalts(knownNumbers, decryptedDict)
    if not salts:
        messagebox.showerror("Ошибка", "Соль не найдена")
        return

    if len(salts) == 1:
        selectedSalt = salts[0]
        resultList = []

        for hashVal, decrypted in decryptedDict.items():
            finalNumber = str(int(decrypted) - selectedSalt)
            resultList.append({"Хеши": hashVal, "Расшифрованные номера": finalNumber})

        resultDf = pd.DataFrame(resultList)

        resultDf.to_excel(phonesFilePath, index=False)

        wb = load_workbook(phonesFilePath)
        ws = wb.active

        ws['C1'] = "Соль"
        ws['C1'].font = Font(bold=True)
        ws['C1'].alignment = Alignment(horizontal="center")

        ws['C2'] = selectedSalt

        ws.column_dimensions['B'].width = len("Расшифрованные номера") + 5

        for column_cells in ws.columns:
            if column_cells[0].column_letter != 'B':
                max_length = 0
                column = [cell for cell in column_cells]
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 4)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save(phonesFilePath)

        messagebox.showinfo("Вывод", f"Результат сохранен в phones.xlsx")
    else:
        with pd.ExcelWriter(phonesFilePath, engine='openpyxl', mode='w') as writer:
            df_salts = pd.DataFrame({"Соли": salts})
            df_salts.to_excel(writer, index=False)

        messagebox.showinfo("Вывод", f"Найдено несколько солей. Они сохранены в phones.xlsx без расшифровки.")

def startDeobfuscationProcess():
    filePath = fileLabel.get().replace("Выбранный для деобезличивания файл: ", "")
    
    decryptedFilePath = runHashcat(filePath)   # Запускаем Hashcat
    
    applyDecryptionSalt(filePath, decryptedFilePath)   # Ищем и применяем соль

def generate_salt(length):
    characters = string.ascii_letters + string.digits
    return ''.join(random.choices(characters, k=length))

def hashPhoneNumbers(HashFunction, salt_type='random'):
    df = pd.read_excel(selectedEncryptionFile)
    phoneNumbers = df.iloc[:, 1].dropna().astype(str).tolist()
    hashedNumbers = []

    salt_length = 1

    if salt_type == 'random':
        salt = generate_salt(salt_length)
    else:
        salt = salt_type

    mask = f"{'?d' * 11}{'?a' * salt_length}"

    hashes_txt_path = os.path.join('hashcat-6.2.6', 'hashes.txt')
    os.makedirs(os.path.dirname(hashes_txt_path), exist_ok=True)

    with open(hashes_txt_path, 'w') as hash_file:
        for number in phoneNumbers:
            salted_number = number + salt

            hash_object = HashFunction(salted_number.encode())
            hashed = hash_object.hexdigest()

            hash_file.write(hashed + '\n')

            hashedNumbers.append({"Хеш": hashed, "Исходный номер": number})

    hashedDf = pd.DataFrame(hashedNumbers)

    algorithm_name = HashFunction.__name__.lower()
    if 'openssl_' in algorithm_name:
        algorithm_name = algorithm_name.replace('openssl_', '')

    hashcat_mapping = {
    'md5': 0,
    'sha1': 100,
    'sha512': 1700,
    }

    hashcat_code = hashcat_mapping.get(algorithm_name, 'unknown')

    hashcat_command = f"hashcat -m {hashcat_code} -a 3 -o found.txt hashes.txt {mask}"
    
    hashedDf.at[0, 'Соль'] = salt
    hashedDf.at[0, 'Код hashcat'] = hashcat_command

    outputFileName = f"phones_{algorithm_name}.xlsx"
    outputFilePath = os.path.join(os.path.dirname(os.path.abspath(__file__)), outputFileName)

    hashedDf.to_excel(outputFilePath, index=False)

    workbook = load_workbook(outputFilePath)
    worksheet = workbook.active

    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    workbook.save(outputFilePath)

    messagebox.showinfo("Вывод", f"Зашифрованные номера сохранены в {outputFileName}")

def hashMD5():
    hashPhoneNumbers(hashlib.md5)

def hashSHA1():
    hashPhoneNumbers(hashlib.sha1)

def hashSHA512():
    hashPhoneNumbers(hashlib.sha512)

def centerWindow(window):
    window.update_idletasks()
    width = window.winfo_width()
    height = window.winfo_height()
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    
    x = (screen_width // 2) - (width // 2)
    y = (screen_height // 2) - (height // 2)
    
    window.geometry(f'{width}x{height}+{x}+{y}')

current_color_index = 0

root = tk.Tk()
root.title("Расшифровка телефонных номеров")

Frame = tk.Frame(root)
Frame.grid(row=0, column=0, padx=10, pady=5, sticky="ew")

fileLabel = tk.StringVar(value="Выбранный для деобезличивания файл: Файл не выбран")
fileLabelDisplay = tk.Label(Frame, textvariable=fileLabel)
fileLabelDisplay.grid(row=0, column=0, padx=0, pady=0, sticky="w")

encryptFileLabel = tk.StringVar(value="Выбранный для зашифровки файл: Файл не выбран")
encryptFileLabelDisplay = tk.Label(Frame, textvariable=encryptFileLabel)
encryptFileLabelDisplay.grid(row=1, column=0, padx=0, pady=0, sticky="w")

buttonFrame_1 = tk.Frame(root)
buttonFrame_1.grid(row=1, column=0, padx=10, pady=5, sticky="ew")

btnSelectFile = tk.Button(buttonFrame_1, text="Выбрать файл для деобезличивания", command=selectEncryptedFile, width=30)
btnSelectFile.grid(row=0, column=0, padx=0, pady=0, sticky="ew")

btnSelectFileForEncryption = tk.Button(buttonFrame_1, text="Выбрать файл для зашифровки", command=selectFileForEncryption, width=34)
btnSelectFileForEncryption.grid(row=0, column=1, padx=4, pady=0, sticky="ew")

buttonFrame_2 = tk.Frame(root)
buttonFrame_2.grid(row=2, column=0, padx=10, pady=5, sticky="ew")

btnDeobfuscate = tk.Button(buttonFrame_2, text="Деобезличить", command=startDeobfuscationProcess, width=30, state=tk.DISABLED)
btnDeobfuscate.grid(row=0, column=0, padx=0, pady=0, sticky="ew")

MD5Button = tk.Button(buttonFrame_2, text="MD5", command=hashMD5, width=10, state=tk.DISABLED)
MD5Button.grid(row=0, column=1, padx=4, pady=0)

sha1Button = tk.Button(buttonFrame_2, text="SHA-1", command=hashSHA1, width=10, state=tk.DISABLED)
sha1Button.grid(row=0, column=2, padx=0, pady=0)

blake2bButton = tk.Button(buttonFrame_2, text="SHA-512", command=hashSHA512, width=10, state=tk.DISABLED)
blake2bButton.grid(row=0, column=3, padx=4, pady=0)

root.update_idletasks()
root.geometry('495x130')
centerWindow(root)

root.mainloop()
