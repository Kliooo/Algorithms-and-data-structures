import tkinter as tk
from tkinter import ttk, filedialog
from openpyxl import load_workbook
from collections import Counter
from datetime import datetime
from dateutil import parser

file_path = None

store_categories = {
    "Магазин электроники": ["М.Видео", "Эльдорадо", "DNS"],
    "Продуктовый магазин": ["Лента", "Пятёрочка", "Ашан"],
    "Розничный магазин": ["Hoff", "OBI", "Леруа Мерлен"]
}

region_mapping = {
    "Центральный район": ["59.934280, 30.335099", "59.934012, 30.304825", "59.934165, 30.325239", 
                          "59.929095, 30.309088", "59.934540, 30.344105", "59.928047, 30.396930", 
                          "59.930199, 30.373353", "59.918032, 30.348554", "59.938720, 30.328290"],
    "Северный район": ["59.956297, 30.364222", "59.958002, 30.371539", "59.944034, 30.384247", 
                       "59.943790, 30.324030", "59.848623, 30.217560", "59.867865, 30.350961", 
                       "59.949219, 30.352710", "59.902220, 30.519794", "59.926789, 30.312345"],
    "Южный район": ["59.845672, 30.251465", "59.850591, 30.252138", "59.845385, 30.323386", 
                    "59.844267, 30.296776", "59.940345, 30.311564", "59.811042, 30.317434", 
                    "59.914567, 30.386810", "59.996297, 30.146508", "59.917654, 30.316543"]
}

category_mapping = {
    "Электроника": ['Смартфоны', 'Ноутбуки', 'Телевизоры', 'Планшеты', 'Наушники'],
    "Бытовая техника": ['Стиральные машины', 'Холодильники'],
    "Здоровое питание": ['Овощи', 'Фрукты','Молочные продукты', 'Кофе и чай'],
    "Менее полезные продукты": ['Хлебобулочные изделия', 'Замороженные продукты', 'Шоколад'],
    "Мебель для гостиной": ['Диваны', 'Столы', 'Стулья'],
    "Мебель для спальни и ванной": ['Кровати', 'Тумбочки', 'Ванны', 'Шкафы']
}

bin_codes = {
    "Сбербанк": ["27402", "27406", "27411", "27416", "27417", "27418", "27420", "27422", "27425", "27427"],
    "ВТБ": ["18868", "18869", "18870", "18873", "21191", "26375", "30127", "31723", "40622", "40623"],
    "ПСБ": ["02507", "04906", "24561", "24562", "24563", "26803", "26804", "29726", "29727", "29728"],
    "Газпромбанк": ["04136", "04270", "04271", "04272", "04273", "24974", "24975", "24976", "26890", "27326"],
    "Альфа-Банк": ["10584", "15400", "15428", "15429", "15481", "15482", "19539", "19540", "21118", "27714"]
}

def setup_window():
    root = tk.Tk()
    root.title("K-anonymity")

    window_width = 480
    window_height = 260

    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    x = (screen_width // 2) - (window_width // 2)
    y = (screen_height // 2) - (window_height // 2)

    root.geometry(f"{window_width}x{window_height}+{x}+{y}")

    return root

def create_widgets(root):
    main_frame = tk.Frame(root)
    main_frame.pack(padx=10, pady=10)

    frame = ttk.LabelFrame(main_frame, text="Выберите квази идентификаторы")
    frame.pack(side='left', padx=10)

    check_vars = {}
    options = ["Название магазина", "Дата и время", "Координаты", "Категория", "Бренд", "Номер карточки", "Количество товаров", "Стоимость"]

    for option in options:
        var = tk.BooleanVar()
        cb = ttk.Checkbutton(frame, text=option, variable=var)
        cb.pack(anchor='w')
        check_vars[option] = var

    k_anonymity_frame = ttk.LabelFrame(main_frame, text="Топ плохих k-anonymity")
    k_anonymity_frame.pack(side='right', padx=10)

    columns = ("Результат",)
    tree = ttk.Treeview(k_anonymity_frame, columns=columns, show='headings', height=7)
    tree.heading("Результат", text="Результат")
    tree.pack()

    button_frame = tk.Frame(root)
    button_frame.pack(pady=10)

    file_button = ttk.Button(button_frame, text="Выбрать файл", command=select_file)
    file_button.pack(side='left', padx=5)

    calculate_button = ttk.Button(button_frame, text="Рассчитать k-anonymity", command=lambda: calculate(tree, check_vars))
    calculate_button.pack(side='left', padx=5)

    anonymize_button = ttk.Button(button_frame, text="Обезличить", command=lambda: anonymize(check_vars))
    anonymize_button.pack(side='left', padx=5)

    exit_button = ttk.Button(button_frame, text="Выход", command=root.quit)
    exit_button.pack(side='left', padx=5)


def select_file():
    global file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        print(f"Выбранный файл: {file_path}")

def calculate(tree, check_vars):
    for item in tree.get_children():
        tree.delete(item)

    try:
        if not file_path:
            print("Файл не выбран!")
            return
        
        workbook = load_workbook(file_path)
        sheet = workbook.active
    except Exception as e:
        print(f"Ошибка чтения файла: {e}")
        return

    headers = [cell.value for cell in sheet[1]]
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    total_records = len(data)
    results = []
    unique_k1_rows = []

    selected_identifiers = [option for option, var in check_vars.items() if var.get()]

    if selected_identifiers:
        selected_indices = [headers.index(identifier) for identifier in selected_identifiers if identifier in headers]

        combinations = [tuple(row[index] for index in selected_indices) for row in data]

        counts = Counter(combinations)
        k_values = list(counts.values())

        rows_to_delete = []
        for i, combination in enumerate(combinations):
            k = counts[combination]
            if k < 10:
                rows_to_delete.append(i + 2)  # Строки в Excel начинаются с 1, +2 из-за заголовка

            if k == 1:
                unique_k1_rows.append(combination)

        num_deleted_rows = len(rows_to_delete)
        percentage_deleted = (num_deleted_rows / total_records) * 100

        for row_idx in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows(row_idx)

        workbook.save(file_path)

        print(f"Удалено {num_deleted_rows} строк из {total_records} ({percentage_deleted:.2f}%).")

        k_values_sorted = sorted(k_values)

        worst_5_k_values = k_values_sorted[:5] if len(k_values_sorted) >= 5 else k_values_sorted

        for k in worst_5_k_values:
            unique_count = k_values.count(k)
            percentage = (k * unique_count / total_records) * 100
            results.append(f"k = {k}   У.к = {unique_count}   {percentage:.2f}%")
    else:
        results.append("Нет квази-идентификатора")

    if unique_k1_rows:
        for row in unique_k1_rows:
            results.append(f"Уникальная строка (k = 1): {row}")

    for result in results:
        tree.insert("", tk.END, values=(result,))

def anonymize(check_vars):
    if not any(var.get() for var in check_vars.values()):
        print("Ошибка: Не выбраны квази-идентификаторы для обезличивания.")
        return
    
    workbook = load_workbook('shopping_dataset.xlsx')
    sheet = workbook.active

    anonymized_workbook = load_workbook('shopping_dataset.xlsx')
    anonymized_sheet = anonymized_workbook.active

    headers = [cell.value for cell in sheet[1]]
    store_index = headers.index("Название магазина")
    time_index = headers.index("Дата и время")
    coordinates_index = headers.index("Координаты")
    category_index = headers.index("Категория")
    brand_index = headers.index("Бренд")
    card_number_index = headers.index("Номер карточки")
    quantity_index = headers.index("Количество товаров")
    cost_index = headers.index("Стоимость")

    anonymization_actions = {
        "Название магазина": lambda row: anonymize_store(row[store_index]),
        "Дата и время": lambda row: anonymize_time(row[time_index]),
        "Координаты": lambda row: anonymize_coordinates(row[coordinates_index]),
        "Категория": lambda row: anonymize_category(row[category_index]),
        "Бренд": lambda row: anonymize_brand(row[brand_index]),
        "Номер карточки": lambda row: anonymize_card(row[card_number_index]),
        "Количество товаров": lambda row: anonymize_quantity(row[quantity_index]),
        "Стоимость": lambda row: anonymize_cost(row[cost_index]),
    }

    selected_identifiers = [identifier for identifier, var in check_vars.items() if var.get()]

    for row in anonymized_sheet.iter_rows(min_row=2, values_only=False):
        for identifier in selected_identifiers:
            if identifier in anonymization_actions:
                anonymization_actions[identifier](row)

    anonymized_workbook.save("anonymized_shopping_data.xlsx")
    print("Данные успешно обезличены и сохранены в anonymized_shopping_data.xlsx")

def anonymize_store(store_cell):
    store_name = store_cell.value
    store_category = None
    for category, stores in store_categories.items():
        if store_name in stores:
            store_category = category
            break

    if store_category:
        store_cell.value = f"{store_category}"

def anonymize_time(time_cell):
    parsed_date = parser.isoparse(time_cell.value)
    month = parsed_date.month
    
    if month in [12, 1, 2]:
        season = "Зима"
    elif month in [3, 4, 5]:
        season = "Весна"
    elif month in [6, 7, 8]:
        season = "Лето"
    else:
        season = "Осень"

    time_cell.value = season

def anonymize_coordinates(coordinates_cell):
    coordinates = coordinates_cell.value
    for region, coords in region_mapping.items():
        if coordinates in coords:
            coordinates_cell.value = region
            break

def anonymize_category(category_cell):
    category_name = category_cell.value
    for general_category, subcategories in category_mapping.items():
        if category_name in subcategories:
            category_cell.value = f"{general_category}"
            break

def anonymize_brand(brand_cell):
    brand_cell.value = "**************"

def anonymize_card(card_number_cell):
    card_number = str(card_number_cell.value)
    if len(card_number) >= 6:
        bin_part = card_number[1:6]
        for bank, bins in bin_codes.items():
            if bin_part in bins:
                card_number_cell.value = bank
                return

def anonymize_quantity(quantity_cell):
    quantity = quantity_cell.value
    if quantity <= 7:
        quantity_cell.value = "5 - 7"
    else:
        quantity_cell.value = "7 - 10"

def anonymize_cost(cost_cell):
    cost = cost_cell.value
    if cost <= 104999:
        cost_cell.value = "1-104999"
    else:
        cost_cell.value = "105000 и более"

if __name__ == "__main__":
    root = setup_window()
    create_widgets(root)
    root.mainloop()
