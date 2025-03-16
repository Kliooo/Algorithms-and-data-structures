import tkinter as tk
from tkinter import ttk
import math
import time
from tkinter.simpledialog import askinteger
from tkinter import filedialog
import tkinter.messagebox as messagebox
from openpyxl import Workbook, load_workbook
import textwrap

global mode
mode = "all"
selected_vertex = None
selected_text_id = None
vertices = []
edges = []
action_stack = []
removed_vertices = []
vertex_counter = 0
min_distance = 30  


def toggle_mode():
    global mode, mode_button
    if mode == "all":
        mode = "one"
        mode_button.config(text="Модификация: Выкл.")
    else:
        mode = "all"
        mode_button.config(text="Модификация: Вкл.")

def add_vertex(event):
    global vertex_counter, removed_vertices
    x, y = event.x, event.y
    
    for vertex in vertices:
        distance = math.sqrt((x - vertex['x'])**2 + (y - vertex['y'])**2)
        if distance < min_distance:
            return

    if removed_vertices:
        vertex_id = min(removed_vertices)
        removed_vertices.remove(vertex_id)
    else:
        vertex_counter += 1
        vertex_id = vertex_counter

    vertex = {'id': vertex_id, 'x': x, 'y': y}
    vertices.append(vertex)

    vertex_oval_id = canvas_input.create_oval(x - 15, y - 15, x + 15, y + 15, fill="black")
    text_id = canvas_input.create_text(x, y, text=str(vertex_id), fill="white")

    action_stack.append(("vertex", vertex, vertex_oval_id, text_id))

def select_vertex_for_edge(event):
    global selected_vertex, selected_text_id
    x, y = event.x, event.y  
    for vertex in vertices:
        distance = math.sqrt((x - vertex['x'])**2 + (y - vertex['y'])**2)
        if distance < 20:  
            if selected_vertex is None:
                selected_vertex = vertex
                selected_text_id = canvas_input.create_text(vertex['x'], vertex['y'] - 28, text="Выбрана вершина: " + str(vertex['id']), fill="green")
                return
            else:
                if selected_vertex != vertex:
                    edge_exists = False
                    edge_index = -1

                    for idx, edge in enumerate(edges):
                        if (edge[0] == selected_vertex['id'] and edge[1] == vertex['id']):
                            edge_exists = True
                            edge_index = idx
                            break

                    if edge_exists:
                        current_weight = edges[edge_index][2]
                        new_weight = askinteger("Вес рёбра", f"Ребро между {selected_vertex['id']} и {vertex['id']} уже существует. Текущий вес: {current_weight}\nВведите новый вес:")
                        if new_weight is not None:
                            old_edge = edges[edge_index]
                            edges[edge_index] = (selected_vertex['id'], vertex['id'], new_weight)
                            tree.item(tree.get_children()[edge_index], values=(selected_vertex['id'], vertex['id'], new_weight))

                            action_stack.append(("weight_change", old_edge, edges[edge_index]))

                        canvas_input.delete(selected_text_id)
                        selected_vertex = None
                        selected_text_id = None
                        return

                    weight = askinteger("Вес рёбра", f"Введите вес рёбра от {selected_vertex['id']} до {vertex['id']}:")
                    if weight is None:
                        return  

                    edge_id = draw_directed_edge(selected_vertex, vertex)
                    if edge_id is not None:
                        edges.append((selected_vertex['id'], vertex['id'], weight, edge_id))
                        tree.insert("", "end", values=(selected_vertex['id'], vertex['id'], weight))     
                        action_stack.append(("edge", (selected_vertex['id'], vertex['id'], weight, edge_id)))

                    canvas_input.delete(selected_text_id)

                    selected_vertex = None  
                    selected_text_id = None  
                    return
                else:
                    selected_vertex = None  
                    canvas_input.delete(selected_text_id)  
                    selected_text_id = None  

def draw_directed_edge(start_vertex, end_vertex):
    radius = 15

    dx = end_vertex['x'] - start_vertex['x']
    dy = end_vertex['y'] - start_vertex['y']

    distance = math.sqrt(dx**2 + dy**2)
    if distance == 0:
        return None

    dx /= distance
    dy /= distance

    start_x = start_vertex['x'] + dx * radius
    start_y = start_vertex['y'] + dy * radius
    end_x = end_vertex['x'] - dx * radius
    end_y = end_vertex['y'] - dy * radius

    edge_id = canvas_input.create_line(start_x, start_y, end_x, end_y, 
                                       arrow=tk.LAST, fill="black", width=3, 
                                       arrowshape=(10, 10, 5))
    return edge_id

def draw_shortest_path(path):
    canvas_output.delete("all")

    for vertex_id in path:
        vertex = next(v for v in vertices if v['id'] == vertex_id)
        canvas_output.create_oval(vertex['x'] - 15, vertex['y'] - 15, vertex['x'] + 15, vertex['y'] + 15, fill="black")
        canvas_output.create_text(vertex['x'], vertex['y'], text=str(vertex['id']), fill="white")

    for i in range(len(path) - 1):
        start_vertex = next(v for v in vertices if v['id'] == path[i])
        end_vertex = next(v for v in vertices if v['id'] == path[i + 1])
        draw_directed_edge_on_canvas(start_vertex, end_vertex)

    start_vertex = next(v for v in vertices if v['id'] == path[-1])
    end_vertex = next(v for v in vertices if v['id'] == path[0])
    draw_directed_edge_on_canvas(start_vertex, end_vertex)

def draw_directed_edge_on_canvas(start_vertex, end_vertex):
    radius = 15

    dx = end_vertex['x'] - start_vertex['x']
    dy = end_vertex['y'] - start_vertex['y']

    distance = math.sqrt(dx**2 + dy**2)
    if distance == 0:
        return None

    dx /= distance
    dy /= distance

    start_x = start_vertex['x'] + dx * radius
    start_y = start_vertex['y'] + dy * radius
    end_x = end_vertex['x'] - dx * radius
    end_y = end_vertex['y'] - dy * radius

    canvas_output.create_line(start_x, start_y, end_x, end_y, 
                               arrow=tk.LAST, fill="red", width=3, 
                               arrowshape=(10, 10, 5))

def calculate_tsp():
    global mode

    for widget in frame_results.winfo_children():
        widget.destroy()

    if len(vertices) < 2:
        results_label = tk.Label(frame_results, text="Недостаточно вершин для решения задачи")
        results_label.pack(padx=0, pady=0, fill="both", expand=True)
        canvas_output.delete("all")
        return

    adj_matrix = {vertex['id']: {} for vertex in vertices}
    for edge in edges:
        adj_matrix[edge[0]][edge[1]] = edge[2]

    start_time = time.time()

    best_path, best_distance = find_best_tsp_path(adj_matrix)

    end_time = time.time()

    elapsed_time = end_time - start_time

    wrapped_best_path = textwrap.fill(f"Лучший путь: {' -> '.join(map(str, best_path))} -> {best_path[0]}", width=100)

    if best_path:
        results_text = (
            f"{wrapped_best_path}\n"
            f"Расстояние: {best_distance}\n"
            f"Время поиска: {elapsed_time:.8f} сек."
        )
        draw_shortest_path(best_path)
    else:
        results_text = "Невозможно найти путь"
        canvas_output.delete("all")

    results_label = tk.Label(frame_results, text=results_text)
    results_label.pack(padx=0, pady=0, fill="both", expand=True)

def find_best_tsp_path(adj_matrix):
    global mode

    best_path = None
    best_distance = float('inf')

    def tsp_from_start(start):
        visited = {v['id']: False for v in vertices}
        current_path = [start['id']]
        current_distance = 0
        visited[start['id']] = True
        current_vertex = start

        while len(current_path) < len(vertices):
            min_dist = float('inf')
            next_vertex = None
            for neighbor_id, weight in adj_matrix[current_vertex['id']].items():
                if not visited[neighbor_id] and weight < min_dist:
                    min_dist = weight
                    next_vertex = neighbor_id

            if next_vertex is None:
                return None, float('inf')

            current_path.append(next_vertex)
            current_distance += min_dist
            visited[next_vertex] = True
            current_vertex = next((v for v in vertices if v['id'] == next_vertex), None)

        if start['id'] in adj_matrix[current_path[-1]]:
            current_distance += adj_matrix[current_path[-1]][start['id']]
        else:
            return None, float('inf')

        return current_path, current_distance

    if mode == "all":
        for start in vertices:
            path, distance = tsp_from_start(start)
            if path and distance < best_distance:
                best_path, best_distance = path, distance

    elif mode == "one":
        start = next((v for v in vertices if v['id'] == 1), None)
        if start:
            best_path, best_distance = tsp_from_start(start)

    return best_path, best_distance

def undo_last_action():
    global selected_vertex, selected_text_id, removed_vertices
    if not action_stack:
        messagebox.showwarning("Нет действий для отмены", "Нет объектов для отмены!")
        return

    last_action = action_stack.pop()

    if last_action[0] == "vertex":
        removed_vertices.append(last_action[1]['id'])
        
        vertices.remove(last_action[1])
        canvas_input.delete(last_action[2])
        canvas_input.delete(last_action[3])

        if selected_vertex == last_action[1]:
            if selected_text_id:
                canvas_input.delete(selected_text_id)
            selected_vertex = None
            selected_text_id = None

    elif last_action[0] == "edge":
        edge_to_remove = last_action[1]
        
        for edge in edges:
            if edge[0] == edge_to_remove[0] and edge[1] == edge_to_remove[1]:
                edges.remove(edge)
                break
        
        if len(edge_to_remove) > 3 and edge_to_remove[3] is not None:
            edge_id = edge_to_remove[3]
            canvas_input.delete(edge_id)

        for item in tree.get_children():
            if tree.item(item, "values")[:2] == (str(edge_to_remove[0]), str(edge_to_remove[1])):
                tree.delete(item)
                break

    elif last_action[0] == "weight_change":
        old_edge, new_edge = last_action[1], last_action[2]

        edge_index = -1
        for idx, edge in enumerate(edges):
            if edge[0] == new_edge[0] and edge[1] == new_edge[1]:
                edge_index = idx
                break

        if edge_index != -1:
            edges[edge_index] = old_edge
            tree.item(tree.get_children()[edge_index], values=(old_edge[0], old_edge[1], old_edge[2]))

def clear_all():
    global vertices, edges, vertex_counter, selected_vertex, selected_text_id
    for item in tree.get_children():
        tree.delete(item)

    canvas_input.delete("all")
    canvas_output.delete("all")

    vertices.clear()
    edges.clear()
    action_stack.clear()
    removed_vertices.clear()

    vertex_counter = 0

    selected_vertex = None
    selected_text_id = None

    for widget in frame_results.winfo_children():
        widget.destroy()

    results_label = tk.Label(frame_results, text="Пусто")
    results_label.pack(padx=0, pady=0, fill="both", expand=True)

def save_graph_to_xlsx():
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return
    
    wb = Workbook()
    ws_vertices = wb.active
    ws_vertices.title = "Vertices"
    ws_vertices.append(["ID", "X", "Y"])
    
    for vertex in vertices:
        ws_vertices.append([vertex['id'], vertex['x'], vertex['y']])
    
    ws_edges = wb.create_sheet(title="Edges")
    ws_edges.append(["Vertex 1", "Vertex 2", "Weight"])
    
    for edge in edges:
        ws_edges.append([edge[0], edge[1], edge[2]])
    
    wb.save(file_path)


def load_graph_from_xlsx():
    global vertex_counter, vertices, edges, action_stack
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return

    clear_all()

    wb = load_workbook(file_path)
    ws_vertices = wb["Vertices"]
    ws_edges = wb["Edges"]

    max_id = 0
    for row in ws_vertices.iter_rows(min_row=2, values_only=True):
        vertex_id, x, y = row
        vertex = {'id': vertex_id, 'x': x, 'y': y}
        vertices.append(vertex)
        max_id = max(max_id, vertex_id)

        vertex_oval_id = canvas_input.create_oval(x - 15, y - 15, x + 15, y + 15, fill="black")
        text_id = canvas_input.create_text(x, y, text=str(vertex_id), fill="white")

        action_stack.append(("vertex", vertex, vertex_oval_id, text_id))

    vertex_counter = max_id

    for row in ws_edges.iter_rows(min_row=2, values_only=True):
        v1, v2, weight = row
        vertex1 = next((v for v in vertices if v['id'] == v1), None)
        vertex2 = next((v for v in vertices if v['id'] == v2), None)
        if vertex1 and vertex2:
            edge_id = draw_directed_edge(vertex1, vertex2)
            edges.append((v1, v2, weight, edge_id))
            tree.insert("", "end", values=(v1, v2, weight))
            
            action_stack.append(("edge", (v1, v2, weight, edge_id)))

def centerWindow(window):
    window.update_idletasks()
    width = window.winfo_width()
    height = window.winfo_height()
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width // 2) - (width // 2)
    y = (screen_height // 2) - (height // 2)
    window.geometry(f'{width}x{height}+{x}+{y}')

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Задача коммивояжера")
    root.geometry("900x600")

    main_frame = tk.Frame(root)
    main_frame.pack(fill="both", expand=True)

    main_frame.grid_columnconfigure(0, weight=1)
    main_frame.grid_columnconfigure(1, weight=1)
    main_frame.grid_rowconfigure(0, weight=1)

    left_frame = tk.Frame(main_frame, padx=0, pady=0)
    left_frame.grid(row=0, column=0, sticky="nsew")

    left_frame.grid_rowconfigure(0, weight=1)
    left_frame.grid_rowconfigure(1, weight=1)
    left_frame.grid_columnconfigure(0, weight=1)

    frame_input = tk.LabelFrame(left_frame, text="Входной граф", padx=0, pady=0)
    frame_input.grid(row=0, column=0, sticky="nsew")

    frame_output = tk.LabelFrame(left_frame, text="Выходной граф", padx=0, pady=0)
    frame_output.grid(row=1, column=0, sticky="nsew")        

    right_frame = tk.Frame(main_frame, padx=0, pady=0)
    right_frame.grid(row=0, column=1, sticky="nsew")

    right_frame.grid_rowconfigure(0, weight=3)
    right_frame.grid_rowconfigure(2, weight=1)
    right_frame.grid_columnconfigure(0, weight=1)

    frame_edges = tk.LabelFrame(right_frame, text="Рёбра", padx=0, pady=0)
    frame_edges.grid(row=0, column=0, sticky="nsew")

    tree = ttk.Treeview(frame_edges, columns=("Vertex 1", "Vertex 2", "Weight"), show="headings")
    tree.heading("Vertex 1", text="Вершина 1")
    tree.heading("Vertex 2", text="Вершина 2")
    tree.heading("Weight", text="Вес")

    tree.pack(fill="both", expand=True)

    tree.column("Vertex 1", width=50, stretch=True)
    tree.column("Vertex 2", width=50, stretch=True)
    tree.column("Weight", width=50, stretch=True)

    frame_buttons = tk.LabelFrame(right_frame, padx=0, pady=0)
    frame_buttons.grid(row=1, column=0, sticky="nsew")

    button_1 = tk.Button(frame_buttons, text="Рассчитать", height=2, command=calculate_tsp)
    button_1.pack(fill="both", expand=True)
    mode_button = tk.Button(frame_buttons, text="Модификация: Вкл.", height=2, command=toggle_mode)
    mode_button.pack(fill="both", expand=True)
    btn_save = tk.Button(frame_buttons, text="Сохранить граф", height=2, command=save_graph_to_xlsx)
    btn_save.pack(fill="both", expand=True)
    btn_load = tk.Button(frame_buttons, text="Загрузить граф", height=2, command=load_graph_from_xlsx)
    btn_load.pack(fill="both", expand=True)
    button_2 = tk.Button(frame_buttons, text="Отменить", height=2, command=undo_last_action)
    button_2.pack(fill="both", expand=True)
    button_3 = tk.Button(frame_buttons, text="Очистить", height=2, command=clear_all)
    button_3.pack(fill="both", expand=True)

    frame_results = tk.LabelFrame(right_frame, text="Результаты", padx=0, pady=0)
    frame_results.grid(row=2, column=0, sticky="ewns")

    results_label = tk.Label(frame_results, text="Пусто")
    results_label.pack(padx=0, pady=0, fill="both", expand=True)

    canvas_input = tk.Canvas(frame_input, bg="white")
    canvas_input.grid(row=1, column=0, sticky="nsew")
    frame_input.grid_rowconfigure(1, weight=1)
    frame_input.grid_columnconfigure(0, weight=1)

    canvas_output = tk.Canvas(frame_output, bg="white")
    canvas_output.grid(row=1, column=0, sticky="nsew")
    frame_output.grid_rowconfigure(1, weight=1)
    frame_output.grid_columnconfigure(0, weight=1)

    canvas_input.bind("<Button-1>", add_vertex)
    canvas_input.bind("<Button-3>", select_vertex_for_edge)

    centerWindow(root)
    root.mainloop()
